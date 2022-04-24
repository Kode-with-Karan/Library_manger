import main
from tkinter import *
import pandas as pd
from tkinter import ttk
from functools import partial
from openpyxl import load_workbook


def show_data():
    df = pd.read_excel("sample.xlsx")
    dataw = Tk()
    dataw.title('Treeview demo')
    screen_width = dataw.winfo_screenwidth()
    screen_height = dataw.winfo_screenheight()
    dataw.geometry(f'620x248+{int(screen_width/3)}+{int(screen_height/4)}')
    wb = load_workbook('sample.xlsx')
    ws = wb.active
    dataw.resizable(False, False)


    def clear_all():
        for item in tree.get_children():
            tree.delete(item)

    def sh_allData():
        clear_all()
        data_arr = main.Lib.showall()
        for i in data_arr:
            tree.insert("", "end", values=i)
        status['text'] = "Showing all data!"

    def sh_bookname():

        def getElement(event):
            selection = event.widget.curselection()
            index = selection[0]
            value = event.widget.get(index)

            clear_all()
            data_arr = main.Lib.sort(['Book Name', value])
            for i in data_arr:
                tree.insert("", "end", values=i)
            top.quit()
            top.destroy()
            status['text'] = "Seccusfully updated from Book Name!"

        top = Tk()
        screen_width = top.winfo_screenwidth()
        screen_height = top.winfo_screenheight()
        top.geometry(f'100x150+{int(screen_width / 3)+45}+{int(screen_height / 4)+50}')

        listbox = Listbox(top)
        for index, cell in enumerate(ws['A']):
            listbox.insert(index, cell.value)
        listbox.pack()
        listbox.bind('<<ListboxSelect>>', getElement)
        top.mainloop()

    def sh_color():

        def getElement(event):
            selection = event.widget.curselection()
            index = selection[0]
            value = event.widget.get(index)

            clear_all()

            data_arr = main.Lib.sort(['Colour', value])
            for i in data_arr:
                tree.insert("", "end", values=i)
            top.quit()
            top.destroy()
            status['text'] = "Seccusfully updated from Colour!"

        top = Tk()
        screen_width = top.winfo_screenwidth()
        screen_height = top.winfo_screenheight()
        top.geometry(f'100x150+{int(screen_width / 3)+45}+{int(screen_height / 4)+50}')
        listbox = Listbox(top)
        for index, cell in enumerate(ws['B']):
            listbox.insert(index, cell.value)
        listbox.pack()
        listbox.bind('<<ListboxSelect>>', getElement)
        top.mainloop()



    def sh_categary():

        def getElement(event):
            selection = event.widget.curselection()
            index = selection[0]
            value = event.widget.get(index)

            clear_all()

            data_arr = main.Lib.sort(['Categary', value])
            for i in data_arr:
                tree.insert("", "end", values=i)
            top.quit()
            top.destroy()
            status['text'] = "Seccusfully updated from Categary!"

        top = Tk()
        screen_width = top.winfo_screenwidth()
        screen_height = top.winfo_screenheight()
        top.geometry(f'100x150+{int(screen_width / 3)+45}+{int(screen_height / 4)+50}')
        listbox = Listbox(top)
        for index, cell in enumerate(ws['C']):
            listbox.insert(index, cell.value)
        listbox.pack()
        listbox.bind('<<ListboxSelect>>', getElement)
        top.mainloop()

    def sh_date():

        def getElement(event):
            selection = event.widget.curselection()
            index = selection[0]
            value = event.widget.get(index)

            clear_all()


            data_arr = main.Lib.sort(['Date', value])

            for i in data_arr:
                tree.insert("", "end", values=i)
            top.quit()
            top.destroy()
            status['text'] = "Seccusfully updated from Data!"

        top = Tk()
        screen_width = top.winfo_screenwidth()
        screen_height = top.winfo_screenheight()
        top.geometry(f'100x150+{int(screen_width / 3)+45}+{int(screen_height / 4)+50}')
        listbox = Listbox(top)
        for index, cell in enumerate(ws['D']):
            listbox.insert(index, cell.value)
        listbox.pack()
        listbox.bind('<<ListboxSelect>>', getElement)
        top.mainloop()

    def sh_time():

        def getElement(event):
            selection = event.widget.curselection()
            index = selection[0]
            value = event.widget.get(index)

            clear_all()

            data_arr = main.Lib.sort(['Time', value])
            for i in data_arr:
                tree.insert("", "end", values=i)
            top.quit()
            top.destroy()
            status['text'] = "Seccusfully updated from Time!"

        top = Tk()
        screen_width = top.winfo_screenwidth()
        screen_height = top.winfo_screenheight()
        top.geometry(f'100x150+{int(screen_width / 3)+45}+{int(screen_height / 4)+50}')
        listbox = Listbox(top)
        for index, cell in enumerate(ws['E']):
            listbox.insert(index, cell.value)
        listbox.pack()
        listbox.bind('<<ListboxSelect>>', getElement)
        top.mainloop()

    del_item = ""
    def sel_deleteItem(event):
        global del_item
        curItem = tree.focus()
        seleted_item = tree.item(curItem)
        del_item = seleted_item['values'][0]

    def deleteItem():
        global del_item
        main.Lib.delete(['Book Name', del_item])
        status['text'] = f"Seccusfully Delete the item {del_item}"
        dataw.destroy()
        show_data()

    m = Menu(dataw)
    dataw.config(menu=m)
    option_menu = Menu(m, tearoff=False)
    m.add_command(label="Back", command=lambda: [dataw.destroy(), main_root()])
    m.add_cascade(label="Sort", menu=option_menu)
    option_menu.add_command(label="All Data", command=sh_allData)
    option_menu.add_command(label="Book Name", command=sh_bookname)
    option_menu.add_command(label="Colour", command=sh_color)
    option_menu.add_command(label="Categary", command=sh_categary)
    option_menu.add_command(label="Date", command=sh_date)
    option_menu.add_command(label="Time", command=sh_time)

    m.add_command(label="Delete", command=deleteItem)
    m.add_command(label="Exit", command=dataw.destroy)

    tree = ttk.Treeview(dataw, height=9)

    # Scroll Bar

    vsb = ttk.Scrollbar(dataw, orient="vertical", command=tree.yview)
    hsb = ttk.Scrollbar(dataw, orient="horizontal", command=tree.xview)
    vsb.grid(column=1,row=0, sticky="NS")
    hsb.grid(column=0, row=1, sticky="EW")

    tree.configure(yscrollcommand=vsb.set)
    tree.configure(xscrollcommand=hsb.set)

    # Data

    tree["columns"] = list(df.columns)
    tree['show'] = 'headings'

    for cols in list(df.columns):
        tree.heading(cols, text=cols)
        tree.column(cols, minwidth=0, width=120)

    for i in df.to_numpy().tolist():
        tree.insert(parent='', index='end', text='', values=(i))


    tree.grid(column=0, row=0)
    tree.bind('<<TreeviewSelect>>', sel_deleteItem)
    status = Label(dataw, text = "Status Bar", bg="black", fg="white")
    status.grid(column=0,row=2, sticky="EW", ipady=2)
    dataw.mainloop()


def main_root():
    def valuest(bookn, color, categary):
        main.Lib.append(bookn.get(),color.get(),categary.get())
        value.set("Successfully Appended!")

    root = Tk()
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    root.geometry(f"300x238+{int(screen_width/2.5)}+{int(screen_height/4)}")

    # Frames

    entryframe = Frame(root)
    entryframe.grid(column=0, row=1, sticky="W", pady=20, padx=20)

    statusframe = Frame(root, width=300,  height=30, bg="black")
    statusframe.grid(sticky='SW')

    # Matter of Menu Farame
    m = Menu(root)
    root.config(menu=m)
    m.add_command(label="Data", command=lambda: [root.destroy(), show_data()])
    m.add_command(label="About", command=valuest)
    m.add_command(label="Help", command=valuest)
    m.add_command(label="Exit", command=root.destroy)

    # Matter of Entryframe

    # Labels

    label = Label(entryframe, text="Book Name : ")
    label.grid(column=0, row=0, ipady=5, ipadx=10, pady=(10, 0))
    label = Label(entryframe, text="Colour : ")
    label.grid(column=0, row=1, ipady=5, ipadx=10, sticky="W")
    label = Label(entryframe, text="Categary : ")
    label.grid(column=0, row=2, ipady=5, ipadx=10, sticky="W")

    # entry Variables
    bookn = StringVar()
    color = StringVar()
    categary = StringVar()
    # entrys
    entry = Entry(entryframe, textvariable=bookn)
    entry.grid(column=1, row=0, pady=(10, 0), padx=(0, 20))
    entry = Entry(entryframe, textvariable=color)
    entry.grid(column=1, row=1, padx=(0, 20))
    entry = Entry(entryframe, textvariable=categary)
    entry.grid(column=1, row=2, padx=(0, 20))

    # Botton
    valuest = partial(valuest, bookn, color, categary)

    btn = Button(entryframe, text="submit", command=valuest)
    btn.grid(column=1, row=3, sticky="W", pady=20)

    # Matter of Status Frame

    # Labels
    value = StringVar()
    value.set("Value")
    label = Label(statusframe, textvariable=value, fg="white", bg="black")
    label.place(relx=0.5, rely=0.5, anchor=CENTER)

    root.mainloop()


if __name__ == '__main__':
    main_root()