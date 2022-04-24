# Library starts here
import os
import datetime
import pandas as pd
from openpyxl import Workbook, load_workbook

if os.path.isfile('sample.xlsx') is True:
    df = pd.read_excel('sample.xlsx')
    wb = load_workbook('sample.xlsx')
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active


ws['A1'] = 'Book Name'
ws['B1'] = "Colour"
ws['C1'] = 'Categary'
ws['D1'] = "Date"
ws['E1'] = "Time"


class Lib:
    @staticmethod
    def append(bname, color, categary):
        date, time = str(datetime.datetime.now()).split(" ")
        # print(bname,  color,  categary)
        ws.append([bname,  color,  categary,  date,  time])
        wb.save("sample.xlsx")

    @staticmethod
    def showall():
        raw_data = []
        for i in df.to_numpy().tolist():
            raw_data.append(i)
        return raw_data

    @staticmethod
    def sort(condition):
        row_data = []
        if condition[0] == 'Book Name':
            for cell in ws['A']:
                if cell.value == condition[1]:
                    print(ws['A'+str(cell.row)].value,
                          ws['B'+str(cell.row)].value,
                          ws['C'+str(cell.row)].value,
                          ws['D'+str(cell.row)].value,
                          ws['E'+str(cell.row)].value)
                    data = (ws['A' + str(cell.row)].value,
                            ws['B' + str(cell.row)].value,
                            ws['C' + str(cell.row)].value,
                            ws['D' + str(cell.row)].value,
                            ws['E' + str(cell.row)].value)
                    row_data.append(data)
            return row_data

        elif condition[0] == 'Colour':
            for cell in ws['B']:
                if cell.value == condition[1]:
                    print(ws['A'+str(cell.row)].value,
                          ws['B'+str(cell.row)].value,
                          ws['C'+str(cell.row)].value,
                          ws['D'+str(cell.row)].value,
                          ws['E'+str(cell.row)].value)
                    data = (ws['A' + str(cell.row)].value,
                            ws['B' + str(cell.row)].value,
                            ws['C' + str(cell.row)].value,
                            ws['D' + str(cell.row)].value,
                            ws['E' + str(cell.row)].value)
                    row_data.append(data)
            return row_data

        elif condition[0] == 'Categary':
            for cell in ws['C']:
                if cell.value == condition[1]:
                    print(ws['A'+str(cell.row)].value,
                          ws['B'+str(cell.row)].value,
                          ws['C'+str(cell.row)].value,
                          ws['D'+str(cell.row)].value,
                          ws['E'+str(cell.row)].value)
                    data = (ws['A'+str(cell.row)].value,
                          ws['B'+str(cell.row)].value,
                          ws['C'+str(cell.row)].value,
                          ws['D'+str(cell.row)].value,
                          ws['E'+str(cell.row)].value)
                    row_data.append(data)
            return row_data

        elif condition[0] == 'Date':
            for cell in ws['D']:
                if cell.value == condition[1]:
                    print(ws['A'+str(cell.row)].value,
                          ws['B'+str(cell.row)].value,
                          ws['C'+str(cell.row)].value,
                          ws['D'+str(cell.row)].value,
                          ws['E'+str(cell.row)].value)
                    data = (ws['A' + str(cell.row)].value,
                            ws['B' + str(cell.row)].value,
                            ws['C' + str(cell.row)].value,
                            ws['D' + str(cell.row)].value,
                            ws['E' + str(cell.row)].value)
                    row_data.append(data)
            return row_data
        elif condition[0] == 'Time':
            for cell in ws['E']:
                if cell.value == condition[1]:
                    print(ws['A'+str(cell.row)].value,
                          ws['B'+str(cell.row)].value,
                          ws['C'+str(cell.row)].value,
                          ws['D'+str(cell.row)].value,
                          ws['E'+str(cell.row)].value)
                    data = (ws['A' + str(cell.row)].value,
                            ws['B' + str(cell.row)].value,
                            ws['C' + str(cell.row)].value,
                            ws['D' + str(cell.row)].value,
                            ws['E' + str(cell.row)].value)
                    row_data.append(data)
            return row_data

    @staticmethod
    def delete(condition):

        if condition[0] == 'Book Name':
            for cell in ws['A']:
                if cell.value == condition[1]:
                    ws.delete_rows(cell.row)
                    wb.save('sample.xlsx')

        if condition[0] == 'Colour':
            for cell in ws['B']:
                if cell.value == condition[1]:
                    ws.delete_rows(cell.row)
                    wb.save('sample.xlsx')

        if condition[0] == 'Categary':
            for cell in ws['C']:
                if cell.value == condition[1]:
                    ws.delete_rows(cell.row)
                    wb.save('sample.xlsx')

        if condition[0] == 'Date':
            for cell in ws['D']:
                if cell.value == condition[1]:
                    ws.delete_rows(cell.row)
                    wb.save('sample.xlsx')

        if condition[0] == 'Time':
            for cell in ws['E']:
                if cell.value == condition[1]:
                    ws.delete_rows(cell.row)
                    wb.save('sample.xlsx')




def data_input():
    bname = input("Enter the Book name: ")
    colour = input("Enter the colour: ")
    categary = input("Enter the Categary: ")

    book = Lib()
    book.append(bname, colour, categary)


if __name__ == '__main__':
    book = Lib()
    print(book.sort(['Date', '2022-04-18']))
